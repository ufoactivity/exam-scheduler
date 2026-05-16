import streamlit as st
import pandas as pd
import numpy as np
import io
import pulp
import traceback
import random
import openpyxl
from datetime import datetime

# ==========================================
# 1. 網頁頁面配置
# ==========================================
st.set_page_config(page_title="段考監考終極自動化", page_icon="🏫", layout="wide")
st.title("🏫 試務組-段考監考全自動化系統 (原格式保留版)")
st.info("💡 視覺大升級：採用 openpyxl 引擎，公布版總表將 100% 完美保留您的原檔案格式（格線、顏色、字體）！")

# --- 初始化狀態 ---
if 'results' not in st.session_state:
    st.session_state['results'] = None
if 'uploader_key' not in st.session_state:
    st.session_state['uploader_key'] = 0

# ==========================================
# 2. 輔助功能定義
# ==========================================
def to_excel_bytes(df, header_df=None):
    output = io.BytesIO()
    if header_df is not None:
        df.columns = header_df.columns
        final_out = pd.concat([header_df, df], ignore_index=True)
    else:
        final_out = df
    
    final_out = final_out.fillna("")
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        final_out.to_excel(writer, index=False, header=False)
    return output.getvalue()

# ==========================================
# 3. 介面佈局
# ==========================================
st.divider()
col1, col2 = st.columns([1, 1], gap="large")

with col1:
    st.subheader("📂 1. 上傳排考資料與範本")
    file_quota = st.file_uploader("1️⃣ 監考堂數.xlsx", type=['xlsx'], key=f"f1_{st.session_state['uploader_key']}")
    file_list = st.file_uploader("2️⃣ 監考名單.xlsx", type=['xlsx'], key=f"f2_{st.session_state['uploader_key']}")
    file_type = st.file_uploader("3️⃣ 監考類型總數.xlsx", type=['xlsx'], key=f"f3_{st.session_state['uploader_key']}")
    file_pub = st.file_uploader("4️⃣ 監考總表公布版.xlsx (範本)", type=['xlsx'], key=f"f4_{st.session_state['uploader_key']}")
    file_assign = st.file_uploader("5️⃣ 監考一覽表.xlsx (班級分配範本)", type=['xlsx'], key=f"f5_{st.session_state['uploader_key']}")

with col2:
    st.subheader("⚙️ 2. 考試設定與日期")
    selected_sheet = None
    if file_quota:
        xls = pd.ExcelFile(file_quota)
        selected_sheet = st.selectbox("👇 選擇考試項目：", xls.sheet_names)
    
    flex_names = []
    if file_list:
        temp_df = pd.read_excel(file_list, header=None).fillna("")
        teacher_list = temp_df.iloc[2:, 1].astype(str).str.strip().tolist()
        teacher_list = [t for t in teacher_list if t != "" and t != "nan"]
        flex_names = st.multiselect("🛡️ 優先時數不大於名單：", options=teacher_list)

    st.write("")
    c_d1, c_d2 = st.columns(2)
    with c_d1: d1_date = st.date_input("📅 第一天日期：", datetime.now())
    with c_d2: d2_date = st.date_input("📅 第二天日期：", datetime.now())
    
    force_run = st.checkbox("⚠️ 忽略健檢警告，強制執行")
    if st.button("🗑️ 清除所有設定", use_container_width=True):
        st.session_state['results'] = None
        st.session_state['uploader_key'] += 1
        st.rerun()

# ==========================================
# 4. 核心演算法執行
# ==========================================
st.divider()

if st.button("🚀 啟動全自動排班與分配", type="primary", use_container_width=True):
    if not all([file_quota, file_list, file_type, file_assign]):
        st.error("🚨 請確認必要檔案皆已上傳！")
    else:
        try:
            df_quota = pd.read_excel(file_quota, sheet_name=selected_sheet).fillna("")
            quota_dict = dict(zip(df_quota.iloc[:, 0].astype(str).str.strip(), pd.to_numeric(df_quota.iloc[:, 1], errors='coerce').fillna(0)))
            
            df_type = pd.read_excel(file_type, header=None).fillna("")
            req_matrix = {'△': [0]*10, '※': [0]*10}
            for i in range(2, len(df_type)):
                row_name = str(df_type.iloc[i, 0]).strip()
                if row_name in ['△', '※']:
                    req_matrix[row_name] = pd.to_numeric(df_type.iloc[i, 1:11], errors='coerce').fillna(0).astype(int).tolist()

            df_list_raw = pd.read_excel(file_list, header=None).fillna("")
            header_df = df_list_raw.iloc[0:2].copy().astype(str).replace('nan', '')
            d1_str, d2_str = d1_date.strftime('%m月%d日'), d2_date.strftime('%m月%d日')
            
            for c in range(3, 8): header_df.iloc[0, c] = d1_str
            for c in range(8, 13): header_df.iloc[0, c] = d2_str
            
            df_list = df_list_raw.iloc[2:].copy()
            teachers = df_list.iloc[:, 1].astype(str).str.strip().tolist()

            # --- PuLP 運算 ---
            with st.spinner("🧠 AI 引擎尋找最佳解中..."):
                prob = pulp.LpProblem("Scheduling", pulp.LpMinimize)
                vX = {}; vY = {}
                for i in range(len(teachers)):
                    vX[i] = {}; vY[i] = {}
                    for j in range(10):
                        vX[i][j] = pulp.LpVariable(f"X_{i}_{j}", cat='Binary')
                        vY[i][j] = pulp.LpVariable(f"Y_{i}_{j}", cat='Binary')
                
                penalty = 0
                for i, t in enumerate(teachers):
                    tgt = int(quota_dict.get(t, 0))
                    act = pulp.lpSum([vX[i][k] + vY[i][k]*2 for k in range(10)])
                    prob += act <= tgt
                    dfct = pulp.LpVariable(f"dfct_{i}", 0)
                    prob += act + dfct == tgt
                    penalty += dfct * (1 if t in flex_names else 1000)
                    for j in range(10):
                        prob += vX[i][j] + vY[i][j] <= 1
                        cell_val = str(df_list.iloc[i, j+3]).strip()
                        if cell_val != "" and cell_val != "nan":
                            prob += vX[i][j] == 0; prob += vY[i][j] == 0
                    prob += vX[i][1] >= vY[i][0]
                    prob += vX[i][6] >= vY[i][5]
                for j in range(10):
                    prob += pulp.lpSum([vX[i][j] for i in range(len(teachers))]) == req_matrix['△'][j]
                    prob += pulp.lpSum([vY[i][j] for i in range(len(teachers))]) == req_matrix['※'][j]
                prob += penalty
                prob.solve()

                schedule_dict = {}
                df_out_master = df_list.copy()
                for i, t in enumerate(teachers):
                    res = []
                    for j in range(10):
                        val = str(df_list.iloc[i, j+3]).strip()
                        if val == "" or val == "nan":
                            if vX[i][j].varValue == 1: val = "△"
                            elif vY[i][j].varValue == 1: val = "※"
                            else: val = "" 
                        res.append(val)
                        df_out_master.iloc[i, j+3] = val
                    schedule_dict[t] = res

            # --- 監考一覽表分配邏輯 ---
            with st.spinner("🎯 執行班級分配..."):
                df_assign_raw = pd.read_excel(file_assign, header=None).fillna("")
                assign_header = df_assign_raw.iloc[0:2].copy().astype(str).replace('nan', '')
                for c in range(1, 6): assign_header.iloc[0, c] = d1_str
                for c in range(6, 11): assign_header.iloc[0, c] = d2_str

                df_assign = df_assign_raw.iloc[2:].copy()
                class_names = df_assign.iloc[:, 0].tolist()
                
                assigned_matrix = np.empty((len(class_names), 10), dtype=object)
                for day_start in [0, 5]:
                    j1 = day_start
                    proctors_j1 = [t for t in teachers if schedule_dict[t][j1] in ["△", "※"]]
                    random.shuffle(proctors_j1)
                    for idx, p in enumerate(proctors_j1): assigned_matrix[idx, j1] = p
                    
                    j2 = day_start + 1
                    proctors_j2 = [t for t in teachers if schedule_dict[t][j2] in ["△", "※"]]
                    bound = {}
                    for idx in range(len(class_names)):
                        p_prev = assigned_matrix[idx, j1]
                        if schedule_dict[p_prev][j1] == "※" and schedule_dict[p_prev][j2] == "△":
                            assigned_matrix[idx, j2] = p_prev
                            bound[p_prev] = True
                    
                    rem = [p for p in proctors_j2 if p not in bound]
                    random.shuffle(rem)
                    r_idx = 0
                    for idx in range(len(class_names)):
                        if assigned_matrix[idx, j2] is None:
                            assigned_matrix[idx, j2] = rem[r_idx]; r_idx += 1

                    for offset in [2, 3, 4]:
                        curr_j = day_start + offset
                        proctors = [t for t in teachers if schedule_dict[t][curr_j] in ["△", "※"]]
                        random.shuffle(proctors)
                        for idx, p in enumerate(proctors): assigned_matrix[idx, curr_j] = p

                for r in range(len(class_names)):
                    for c in range(10): df_assign.iloc[r, c+1] = assigned_matrix[r, c]

            # --- 公布版套印 (Openpyxl 完美保留格式) ---
            pub_bytes = None
            if file_pub:
                with st.spinner("🖨️ 正在將資料無縫套印至公布版..."):
                    # 讀取 Excel 並保留所有樣式與格式
                    wb = openpyxl.load_workbook(file_pub)
                    ws = wb.active
                    
                    h_row = -1
                    t_cols = []
                    
                    # 搜尋 "教師" 欄位 (openpyxl 的索引是從 1 開始的)
                    for r in range(1, min(20, ws.max_row + 1)):
                        for c in range(1, ws.max_column + 1):
                            val = ws.cell(row=r, column=c).value
                            if val and "教師" in str(val):
                                h_row = r
                                t_cols.append(c)
                        if h_row != -1:
                            break
                    
                    if h_row != -1:
                        for c in t_cols:
                            # 填寫日期
                            if h_row - 1 >= 1:
                                ws.cell(row=h_row-1, column=c+2).value = d1_str
                                ws.cell(row=h_row-1, column=c+7).value = d2_str
                            
                            # 向下尋找老師並填寫排班
                            for r in range(h_row+1, ws.max_row + 1):
                                t_val = ws.cell(row=r, column=c).value
                                if t_val is not None:
                                    name = str(t_val).strip()
                                    if name in schedule_dict:
                                        # 第一天 (5節)
                                        for j in range(5): 
                                            ws.cell(row=r, column=c+2+j).value = schedule_dict[name][j]
                                        # 第二天 (5節)
                                        for j in range(5): 
                                            ws.cell(row=r, column=c+7+j).value = schedule_dict[name][j+5]
                    
                    # 將修改後的 Workbook 存入 BytesIO
                    output = io.BytesIO()
                    wb.save(output)
                    pub_bytes = output.getvalue()

            st.balloons()
            st.session_state['results'] = {
                'orig': to_excel_bytes(df_out_master, header_df),
                'assign': to_excel_bytes(df_assign, assign_header),
                'pub': pub_bytes
            }

        except Exception as e:
            st.error(f"發生錯誤: {e}")
            st.code(traceback.format_exc())

# ==========================================
# 5. 下載區
# ==========================================
if st.session_state['results']:
    st.divider()
    res = st.session_state['results']
    c1, c2, c3 = st.columns(3)
    with c1: st.download_button("📥 1. 監考總表", res['orig'], "監考總表.xlsx", "application/vnd.ms-excel", use_container_width=True)
    with c2: st.download_button("📥 2. 監考一覽表(分配版)", res['assign'], "監考一覽表_分配完成.xlsx", "application/vnd.ms-excel", use_container_width=True, type="primary")
    with c3: 
        if res['pub']: st.download_button("📥 3. 公布版套印總表", res['pub'], "公布版總表.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
